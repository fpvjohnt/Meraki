import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from src.utils.logger import logger
from typing import Dict, Any

def generate_excel_report(results: Dict[str, Any], org_name: str):
    logger.info(f"Generating Excel report for organization: {org_name}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Health Check Results"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Write headers
    headers = ["Network", "Check", "Status", "Details"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for network, checks in results.items():
        if network == "org_settings":
            network_name = "Organization Settings"
        else:
            network_name = network

        for check, result in checks.items():
            ws.cell(row=row, column=1, value=network_name)
            ws.cell(row=row, column=2, value=check)
            
            if isinstance(result, dict) and 'is_ok' in result:
                status = "OK" if result['is_ok'] else "Issue"
                ws.cell(row=row, column=3, value=status)
                
                details = ""
                if 'issues' in result:
                    for issue in result['issues']:
                        details += f"- {issue['message']}\n"
                elif 'message' in result:
                    details = result['message']
                
                ws.cell(row=row, column=4, value=details)
            else:
                ws.cell(row=row, column=3, value="N/A")
                ws.cell(row=row, column=4, value=str(result))

            row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    filename = f"{org_name}_health_check_report.xlsx"
    wb.save(filename)
    logger.info(f"Excel report saved as {filename}")